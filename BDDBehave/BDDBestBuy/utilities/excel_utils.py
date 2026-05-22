import openpyxl


class ExcelUtils:

    # ==========================================================
    # READ VALID BRAND FILTER DATA
    # ==========================================================
    @staticmethod
    def get_brand_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["BrandFilters"]

        data = []

        for row in range(
                2,
                sheet.max_row + 1
        ):

            brand = sheet.cell(
                row,
                1
            ).value

            data.append(brand)

        return data

    # ==========================================================
    # READ VALID PRICE FILTER DATA
    # ==========================================================
    @staticmethod
    def get_price_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["PriceFilters"]

        return (
            sheet.cell(2, 1).value,
            sheet.cell(2, 2).value
        )

    # ==========================================================
    # READ INVALID PRICE DATA
    # ==========================================================
    @staticmethod
    def get_invalid_price_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook[
            "InvalidPriceFilters"
        ]

        data = []

        for row in range(
                2,
                sheet.max_row + 1
        ):

            min_price = sheet.cell(
                row,
                1
            ).value

            max_price = sheet.cell(
                row,
                2
            ).value

            data.append(
                (
                    str(min_price),
                    str(max_price)
                )
            )

        return data

    # ==========================================================
    # READ INVALID BRAND DATA
    # ==========================================================
    @staticmethod
    def get_invalid_brand_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook[
            "InvalidBrandFilter"
        ]

        return sheet.cell(2, 1).value

    # ==========================================================
    # READ INVALID EMAIL DATA
    # ==========================================================
    @staticmethod
    def get_invalid_email_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["InvalidEmail"]

        return sheet.cell(2, 1).value