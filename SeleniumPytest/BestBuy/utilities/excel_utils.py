import openpyxl


class ExcelUtils:

    @staticmethod
    def get_brand_data():
        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["BrandFilters"]

        data = []

        # Read all brand names
        for row in range(2, sheet.max_row + 1):
            brand_name = sheet.cell(row, 1).value

            data.append(brand_name)

        return data

    @staticmethod
    def get_price_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["PriceFilters"]

        return (
            sheet.cell(2, 1).value,
            sheet.cell(2, 2).value
        )

    @staticmethod
    def get_invalid_price_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["InvalidPriceFilters"]

        data = []

        # Read all invalid price ranges
        for row in range(2, sheet.max_row + 1):
            min_price = sheet.cell(row, 1).value
            max_price = sheet.cell(row, 2).value

            data.append(
                (
                    str(min_price),
                    str(max_price)
                )
            )

        return data

    @staticmethod
    def get_invalid_brand_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["InvalidBrandFilter"]

        return sheet.cell(2, 1).value

    @staticmethod
    def get_invalid_email_data():

        workbook = openpyxl.load_workbook(
            "test_data/test_data.xlsx"
        )

        sheet = workbook["InvalidEmail"]

        return sheet.cell(2, 1).value